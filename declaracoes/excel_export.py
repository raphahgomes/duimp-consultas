"""
Geração do arquivo Excel com os itens de uma declaração.

Produz arquivo .xlsx com as mesmas 12 colunas de BOTDI.py:
  Nº Adição | Seq. | Quantidade | Unidade Medida | Descrição Mercadoria |
  cClassTrib | Valor Unitário | II | IPI | PIS/PASEP | COFINS | NCM
"""

import logging
import os
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from declaracoes.models import ConsultaDeclaracao

logger = logging.getLogger(__name__)

COLUNAS = [
    ('Nº Adição',          8),
    ('Seq.',               6),
    ('Quantidade',        16),
    ('Unidade Medida',    14),
    ('Descrição Mercadoria', 60),
    ('cClassTrib',        12),
    ('Valor Unitário',    18),
    ('II',                 8),
    ('IPI',                8),
    ('PIS/PASEP',         10),
    ('COFINS',            10),
    ('NCM',               12),
]

# Cores do cabeçalho (azul escuro, texto branco)
HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
DATA_FONT = Font(name='Calibri', size=10)


def gerar_excel(consulta: 'ConsultaDeclaracao', destino: str | Path) -> Path:
    """
    Gera o arquivo Excel para a consulta e salva em `destino`.

    Parâmetros
    ----------
    consulta : ConsultaDeclaracao com itens já salvos no banco
    destino  : caminho completo do arquivo a criar (incluindo .xlsx)

    Retorna
    -------
    Path do arquivo gerado.
    """
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = consulta.numero[:31]  # Limite de 31 caracteres para nomes de abas

    # ── Cabeçalho ──────────────────────────────────────────────────────────
    for col_idx, (titulo, largura) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.row_dimensions[1].height = 30

    # ── Dados ──────────────────────────────────────────────────────────────
    itens = consulta.itens.order_by('num_adicao', 'sequencial')
    for row_idx, item in enumerate(itens, start=2):
        valores = [
            item.num_adicao,
            item.sequencial,
            item.quantidade,
            item.unidade_medida,
            item.descricao,
            item.c_class_trib,
            item.valor_unitario,
            item.ii,
            item.ipi,
            item.pis_pasep,
            item.cofins,
            item.ncm,
        ]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx == 5))

        # Zebrado: linhas pares em cinza claro
        if row_idx % 2 == 0:
            for col_idx in range(1, len(COLUNAS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill('solid', fgColor='F2F2F2')

    # ── Auto-filtro ────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Congela cabeçalho ──────────────────────────────────────────────────
    ws.freeze_panes = 'A2'

    wb.save(destino)
    logger.info('Excel gerado: %s (%d itens)', destino, itens.count())
    return destino
